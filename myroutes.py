from flask import  request,  send_file, render_template, redirect, jsonify, url_for
from openpyxl import load_workbook, Workbook
  

def submit_file():
    wb = load_workbook('RegisteredUsers.xlsx') 
    sheet1 = wb.active  
    top=sheet1.max_row+1
    f = request.files['pp']
    if(f.filename!=''):
        f.save('static/files/'+f.filename)

        sheet1["A"+str(top)] = request.form['name']
        sheet1["B"+str(top)] = request.form['email']  
        sheet1["C"+str(top)] = request.form['password'] 
        sheet1["D"+str(top)] = f.filename 
        print(f.filename)
        wb.save('RegisteredUsers.xlsx') 
        return redirect('/login')
    else:
        print("No file choosen")
        return redirect('/register')
    

def home():
    return render_template('home.html')

def register():
    return render_template('register.html')

def login():
    return render_template('login.html')



def sendlogdata(user_name, type):
    if(type == "xlsx"):
        wb2 = load_workbook('Logindata.xlsx')
        loginsheet = wb2.active 
        tempwb = Workbook()
        tempsheet = tempwb.active
        tempsheettop = 1
        tempsheet["A"+str(tempsheettop)] = "Date"
        tempsheet["B"+str(tempsheettop)] = "Time"

        for i in range(2, loginsheet.max_row):
            if(loginsheet["A"+str(i)].value == user_name ):
                tempsheet["A"+str(tempsheettop+1)] = loginsheet["B"+str(i)].value
                tempsheet["B"+str(tempsheettop+1)] = loginsheet["C"+str(i)].value
                tempsheettop = tempsheettop+1

        tempwb.save(user_name+".xlsx")
        return send_file(user_name+".xlsx", cache_timeout=0)
    else:
        returnthis = {}
        wb2 = load_workbook('Logindata.xlsx')
        loginsheet = wb2.active

        top=1
        for i in range(2, loginsheet.max_row+1):
            if(loginsheet["A"+str(i)].value == user_name):
                temp = {"Date":loginsheet["B"+str(top)].value , "Time":loginsheet["C"+str(top)].value }
                returnthis[str(top)] = temp
                top =top+1
        return jsonify(returnthis)
    
def logLogin(user_name):
    wb2 = load_workbook('Logindata.xlsx')
    loginsheet = wb2.active
    top = loginsheet.max_row+1
    loginsheet["A"+str(top)] = user_name
    from datetime import datetime
    time = datetime.now().strftime("%H:%M:%S")
    date = datetime.now().strftime("%d/%m/%y")
    loginsheet["B"+str(top)] = date
    loginsheet["C"+str(top)] = time
    wb2.save('LoginData.xlsx')

def profile(user_name, img_file_path):
    #saving time and date data to file
    logLogin(user_name)
    img_path ="\\static\\files\\"+img_file_path
    return render_template('profile_page.html', name=user_name, src=img_path)


def loginAuth():
    wb = load_workbook('RegisteredUsers.xlsx') 
    sheet1 = wb.active 
    email = request.form["email"]
    password = request.form["password"] 
    for row in range(2, sheet1.max_row+1):
        if(sheet1['B'+str(row)].value==email):
            if(sheet1['C' + str(row)].value == password):
                return redirect(url_for('profile', user_name =sheet1['A' + str(row)].value, img_file_path =sheet1['D' + str(row)].value) )
            
    return jsonify({'status':'login failed'})

# $env:FLASK_APP = "helloflask.py"
# $env:FLASK_ENV = "development"
